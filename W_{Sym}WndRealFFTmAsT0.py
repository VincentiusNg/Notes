import openpyxl
from openpyxl.cell import get_column_letter, column_index_from_string

wb = openpyxl.load_workbook('Radar_RDM_PCL_FFT_Integrated_Windows.xlsx', data_only=True)
#Debug: list all sheet on workbook
wb.get_sheet_names()
#
sheet = wb.get_sheet_by_name('W_{Sym}WndRealFFTmAsT0')
#Debug: print name and title of current sheet
sheet
sheet.title

"""
typedef struct
{
	RAI_U16 	tcId = 0
	char		tcName[128] = 0 // Reserve purpose
	char		inFile[128] = 0
	char		outFile[128] = 0
	RAI_U32		inSize = 0
	RAI_U32		outSize = 0
	RAI_U16		win_fft = 0	//0: No processing | 1: Windows
	RAI_U16 	hsym_sym = 0	//0: None | 1: hsym | 2: sym
	RAI_U16		cplx_real = 0	//0: Complex | 1: Real
	RAI_U16		point = 0
	RAI_U16		scale = 0		//0: As | 1: Us | 2: Fs
	RAI_U16		coeff = 0		//0: None | 1: rectangle(default) | 2: T.B.D
	RAI_S32		block_exp = 0	//value
	RAI_S32		f_scale = 0	//value
	RAI_S32		r_scale = 0	//value T.B.D
	char		padding[92] = 0	//
} param_fft_t = 0

    [13] =
    {
    		13,
    		"RDM_DSP_SymWndRealFFT512AsT0",
    		"dataRealFFT512AsT0.in",
    		"dataSymWRealFFT512AsT0_outN",
    		16384,	//16 Kb
    		8192,	//8 Kb
    		WINDOW,
    		SYM,
    		REAL,
    		POINT512,
    		AS,
    		NON_COEFF,
    		0,
    		0,
    		0
    },
"""
"""
Idea:
"""
tcId = 0
tcName = ""		#OK
inFile = ""		#OK
outFile = ""	     #OK
inSize = 0
outSize = 0
win_fft = 0	     #cmp
hsym_sym = 0	     #cmp
cplx_real = 0       #cmp	
point = 0		#OK
scale = 0		
coeff = 0		#OK	
block_exp = 0	#OK
f_scale = 0
r_scale = 0		#OK


#range of column
col_start = column_index_from_string('E')
col_stop = column_index_from_string('U')
#debug
#col_stop = column_index_from_string('G')

#range of row
row_start = 6
row_stop = 89
#range of API
row_start_API = 7
row_stop_API  = 22
#range of point
row_start_point = 23
row_stop_point  = 32
#range of dataInput
row_start_dataInput = 33
row_stop_dataInput  = 43
#rang of coefficient
row_start_coeff = 44
row_stop_coeff  = 46
#rang of block_exp
row_start_block_exp = 47
row_stop_block_exp  = 49
#range of Twdtable
#row_start_twd_tbl = 50
#row_stop_twd_tbl  = 54
#range of dataOutput
row_start_dataOutput1 = 55
row_stop_dataOutput1 = 64
row_start_dataOutput2 = 75
row_stop_dataOutput2 = 80
#range of r_scaling
row_start_r_scaling1 = 65
row_stop_r_scaling1 = 74
row_start_r_scaling2 = 81
row_stop_r_scaling2 = 86

resultFile = open('result.txt', 'w')
for icol in range(col_start, col_stop):
    tcId = sheet.cell(row=row_start, column=icol).value
    print('tcId', tcId) 
    for irow in range(row_start, row_stop):
        if(sheet.cell(row=irow, column=icol).value) == 'o':
#            print("dbg checking...", "irow = ", irow, "icol = ", get_column_letter(icol))
            tmp = sheet.cell(row=irow, column=column_index_from_string('D')).value
            if((irow >= row_start_API) and (irow <= row_stop_API)):
                tcName = tmp
                print("API: ", tcName)
            # extract win_fft, hsym_sym, cplx_real
                if "Wnd" in tcName:
                    print("win_fft", " WINDOW")
                    win_fft = 1
                    if "HSym" in tcName:
                        print("hsym_sym", " HSYM")
                        hsym_sym = 1
                    elif "Sym" in tcName:
                        print("hsym_sym", " SYM")
                        hsym_sym = 2
                    else:
                        print("hsym_sym", " NO_HSYM_SYM")
                        hsym_sym = 0    
                elif "SpctConv" in tcName:
                    print("win_fft", " SPECTRUM_CONV")
                    win_fft = 2
                else:
                    print("win_fft", " NO_PROCESSING")
                    win_fft = 0;
            # extract cplx_real    
                if "Cplx" in tcName:
                    print ("cplx_real"," COMPLEX")
                    cplx_real = 0
                elif "Real" in tcName:
                    print ("cplx_real"," REAL")
                    cplx_real = 1
                else:
                    print("the tc not belong to cplx or real")
            #extract scale
                if "As" in tcName:
                    print ("scale"," AS")                    
                    scale = 0
                elif "Us" in tcName:
                    print ("scale"," US")                    
                    scale = 1                    
                elif "Fs" in tcName:
                    print ("scale"," FS")                    
                    scale = 2
                else:
                    print("the tc not belong to scale")
            if((irow >= row_start_point) and (irow <= row_stop_point)):
                point = tmp
                print("point: ", point)
            if((irow >= row_start_dataInput) and (irow <= row_stop_dataInput)):
                inFile = tmp
                print("dataInput: ", inFile)
            if((irow >= row_start_coeff) and (irow <= row_stop_coeff)):
                coeff = tmp
                print("coeff: ", coeff)
            if((irow >= row_start_block_exp) and (irow <= row_stop_block_exp)):
                block_exp = tmp
                print("block_exp: ", block_exp)
#                print("dbg checking...", "irow = ", irow, "icol = ", get_column_letter(icol))
            if(((irow >= row_start_dataOutput1) and (irow <= row_stop_dataOutput1)) or ((irow >= row_start_dataOutput2) and (irow <= row_stop_dataOutput2))):
                outFile = tmp
                print("outFile: ", outFile)
            if(((irow >= row_start_r_scaling1) and (irow <= row_stop_r_scaling1)) or ((irow >= row_start_r_scaling2) and (irow <= row_stop_r_scaling2))):
                r_scale = tmp
                print("r_scale: ", r_scale)				
        else:
            print ("Nothing")
#            print(tmp)          
    print('--- END OF COLUMN ---')
    print('Writing results...')
    
    resultFile.write('\t' + '['+ str(tcId) +']'+' = \n')  
    resultFile.write('\t' + '{' +'\n')  
    resultFile.write('\t\t' + str(tcId) + ','+'\n')
    resultFile.write('\t\t' + '"' + str(tcName).strip()+ '"' + ','+'\n')
    resultFile.write('\t\t' + '"' + str(inFile).strip()+ '"' + ','+'\n')
    resultFile.write('\t\t' + '"' + str(outFile).strip()+ '"' + ','+'\n')
    #T.B.D
    inSize = point * 4;
    outSize = ((point * 4)/2)+1;
    resultFile.write('\t\t' + str(inSize) + ','+'\n')
    resultFile.write('\t\t' + str(outSize) + ','+'\n')
    #win_fft
    if(win_fft == 0):
        resultFile.write('\t\t' + 'NO_PROCESSING' + ','+'\n')
    elif(win_fft == 1):
        resultFile.write('\t\t' + 'WINDOW' + ','+'\n')
    elif(win_fft == 2):
        resultFile.write('\t\t' + 'SPECTRUM_CONV' + ','+'\n')
    else:
        resultFile.write('\t\t' + 'N/A' + ','+'\n')
    #hsym_sym
    if(hsym_sym == 0):
        resultFile.write('\t\t' + 'NO_HSYM_SYM' + ','+'\n')
    elif(hsym_sym == 1):
        resultFile.write('\t\t' + 'HSYM' + ','+'\n')
    elif(hsym_sym == 2):
        resultFile.write('\t\t' + 'SYM' + ','+'\n')
    else:
        resultFile.write('\t\t' + 'N/A' + ','+'\n')
    #cplx_real
    if(cplx_real == 0):
        resultFile.write('\t\t' + 'COMPLEX' + ','+'\n')
    elif(cplx_real == 1):
        resultFile.write('\t\t' + 'REAL' + ','+'\n')
    else:
        resultFile.write('\t\t' + 'N/A' + ','+'\n')
    #point
    resultFile.write('\t\t' + 'POINT'+str(point) + ','+'\n')
    #scale
    if(scale == 0):
        resultFile.write('\t\t' + 'AS' + ','+'\n')
    elif(scale == 1):
        resultFile.write('\t\t' + 'US' + ','+'\n')
    elif(scale == 2):
        resultFile.write('\t\t' + 'FS' + ','+'\n')
    else:
        resultFile.write('\t\t' + 'N/A' + ','+'\n')
    #window coeff (rectagle == 1)
    if "Rectangle" in str(coeff):
        resultFile.write('\t\t' + 'RECTANGLE' + ','+'\n')
    else:
        resultFile.write('\t\t' + 'NON_COEFF' + ','+'\n')
    #block_exp
    resultFile.write('\t\t' + str(block_exp) + ','+'\n')    
    #final scaling
    resultFile.write('\t\t' + str(f_scale) + ','+'\n')
    #T.B.D r_scale
#    resultFile.write('\t\t' + str(r_scale) + ','+'\n')
    resultFile.write('\t\t' + '0' + ','+'\n')
    
    resultFile.write('\t' + '}' + ','+'\n')    
    
    #reset value
    tcId = 0
    tcName = ""		#OK
    inFile = ""		#OK
    outFile = ""	     #OK
    inSize = 0
    outSize = 0
    win_fft = 0	     #cmp
    hsym_sym = 0	     #cmp
    cplx_real = 0       #cmp	
    point = 0		#OK
    scale = 0		
    coeff = 0		#OK	
    block_exp = 0	#OK
    f_scale = 0
    r_scale = 0		#OK
resultFile.close()
print('Done.')		