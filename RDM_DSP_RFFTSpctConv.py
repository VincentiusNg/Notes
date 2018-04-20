import openpyxl
from openpyxl.cell import get_column_letter, column_index_from_string

wb = openpyxl.load_workbook('Radar_RDM_PCL_FFT_Spectrum_conversion.xlsx', data_only=True)
#Debug: list all sheet on workbook
wb.get_sheet_names()
#
sheet = wb.get_sheet_by_name('Spectrum_Conversion')
#Debug: print name and title of current sheet
sheet
sheet.title

"""
Initialize value param_fft_t parameter
"""
tcId = 0
tcName = ""		
inFile = ""	
outFile = ""	     
inSize = 0
outSize = 0
win_fft = 0	     
hsym_sym = 0	     
cplx_real = 0       
point = 0		
scale = 0		
coeff = 0		
block_exp = 0	
f_scale = 0
r_scale = 0		


#range of column (D -> P)
col_start = column_index_from_string('D')
col_stop = column_index_from_string('P')

#range of row
row_start = 6
row_stop = 81
#range of API
#row_start_API = 0   #skip
#row_stop_API  = 22  #skip
#range of point
row_start_point = 8
row_stop_point  = 17
#range of dataInput
row_start_dataInput = 19
row_stop_dataInput  = 28
#rang of coefficient
#row_start_coeff = 44    #skip
#row_stop_coeff  = 46    #skip    
#rang of block_exp
#row_start_block_exp = 47 #skip
#row_stop_block_exp  = 49 #skip

#range of dataOutput
row_start_dataOutput1 = 58
row_stop_dataOutput1 = 67
#range of r_scaling
row_start_r_scaling1 = 68
row_stop_r_scaling1 = 77

resultFile = open('RDM_DSP_RFFTSpctConvList.txt', 'w')
resultFile.write('static param_fft_t RDM_DSP_RFFTSpctConvList[TC_RDM_DSP_RFFTSPCTCONV] =\n')
resultFile.write('{\n')
for icol in range(col_start, col_stop):
    tcId = sheet.cell(row=row_start, column=icol).value
    print('tcId', tcId) 
    for irow in range(row_start, row_stop):
        if((sheet.cell(row=irow, column=icol).value) == 'o') or ((sheet.cell(row=irow, column=icol).value) == '-'):
            print("dbg checking...", "irow = ", irow, "icol = ", get_column_letter(icol))
            tmp = sheet.cell(row=irow, column=column_index_from_string('C')).value
            
            tcName = "RDM_DSP_RFFTSpctConv"
            print("API: ", tcName)
            # extract win_fft, hsym_sym, cplx_real
            hsym_sym = 0    
            win_fft = 2;    #fixed value for SPECTRUM CONV
            cplx_real = 0   #fixed value "0"            
            scale = 0
            if((irow >= row_start_point) and (irow <= row_stop_point)):
                point = tmp
                print("point: ", point)
            if((irow >= row_start_dataInput) and (irow <= row_stop_dataInput)):
                inFile = tmp
                print("dataInput: ", inFile)
            coeff = 0
            print("coeff: ", coeff)
            block_exp = 0
            print("block_exp: ", block_exp)
            print("dbg checking...", "irow = ", irow, "icol = ", get_column_letter(icol))
            if((irow >= row_start_dataOutput1) and (irow <= row_stop_dataOutput1)):
                outFile = tmp
                print("outFile: ", outFile)
            if((irow >= row_start_r_scaling1) and (irow <= row_stop_r_scaling1)):
                r_scale = tmp
                print("r_scale: ", r_scale)				
        else:
            print ("Nothing")
#            print(tmp)          
    print('--- END OF COLUMN ---')
    print('Writing results...')
 
    resultFile.write('\t' + '{' +'\n')  
    resultFile.write('\t\t' + str(tcId) + ','+'\n')
    resultFile.write('\t\t' + '"' + str(tcName).strip()+ '"' + ','+'\n')
    resultFile.write('\t\t' + '"' + str(inFile).strip()+ '"' + ','+'\n')
    resultFile.write('\t\t' + '"' + str(outFile).strip()+ '"' + ','+'\n')
    #T.B.D
    inSize = ((point * 4)/2)+1;
    outSize = point * 4;
    resultFile.write('\t\t' + str(inSize) + ','+'\n')
    resultFile.write('\t\t' + str(outSize) + ','+'\n')
    #win_fft
    if(win_fft == 0):
        resultFile.write('\t\t' + 'NO_PROCESSING' + ','+'\n')
    elif(win_fft == 1):
        resultFile.write('\t\t' + 'WINDOW' + ','+'\n')
    elif(win_fft == 2):
        resultFile.write('\t\t' + 'SPECTRUM_CONV' + ','+'\n')
    else:
        resultFile.write('\t\t' + 'N/A' + ','+'\n')
    #hsym_sym
    if(hsym_sym == 0):
        resultFile.write('\t\t' + 'NO_HSYM_SYM' + ','+'\n')
    elif(hsym_sym == 1):
        resultFile.write('\t\t' + 'HSYM' + ','+'\n')
    elif(hsym_sym == 2):
        resultFile.write('\t\t' + 'SYM' + ','+'\n')
    else:
        resultFile.write('\t\t' + 'N/A' + ','+'\n')
    #cplx_real
    resultFile.write('\t\t' + '0' + ','+'\n')

    #point
    resultFile.write('\t\t' + 'POINT'+str(point) + ','+'\n')
    #scale
    resultFile.write('\t\t' + '0' + ','+'\n')
    #window coeff (rectagle == 1)
    if "Rectangle" in str(coeff):
        resultFile.write('\t\t' + 'RECTANGLE' + ','+'\n')
    else:
        resultFile.write('\t\t' + 'NON_COEFF' + ','+'\n')
    #block_exp
    resultFile.write('\t\t' + str(block_exp) + ','+'\n')    
    #final scaling
    resultFile.write('\t\t' + str(f_scale) + ','+'\n')
    #T.B.D r_scale
#    resultFile.write('\t\t' + str(r_scale) + ','+'\n')
    resultFile.write('\t\t' + '0' + ','+'\n')
    
    resultFile.write('\t' + '}' + ','+'\n')    
    
    #reset value
    tcId = 0
    tcName = ""		#OK
    inFile = ""		#OK
    outFile = ""	     #OK
    inSize = 0
    outSize = 0
    win_fft = 0	     #cmp
    hsym_sym = 0	     #cmp
    cplx_real = 0       #cmp	
    point = 0		#OK
    scale = 0		
    coeff = 0		#OK	
    block_exp = 0	#OK
    f_scale = 0
    r_scale = 0		#OK
resultFile.write('};\n')
resultFile.close()
print('Done.')		