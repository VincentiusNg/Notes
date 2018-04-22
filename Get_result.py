import math
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

wb = openpyxl.load_workbook('result.xlsx', data_only = True)
#sheetList = wb.get_sheet_names()
sheetList = wb.sheetnames
#sheet = wb.get_sheet_by_name('result2')
workSheet = wb[sheetList[0]]

rowStart = 3
binOutCol = column_index_from_string('B')
bxpCol = column_index_from_string('F')
bxpOutCol = column_index_from_string('G')

maxErrorCol = column_index_from_string('H')
maxErrorRateCol = column_index_from_string('K')

countErrorRate1Col = column_index_from_string('N')
countErrorRate2Col = column_index_from_string('Q')
countErrorRate3Col = column_index_from_string('T')

constJudge1 = 0.001
constJudge2 = 0.05
constJudgeMaxError = 1
constJudgeBxp = 0.015

#Get number of testcases
rowEnd = rowStart
numberOfTestcases = 0
passed = 0;
failed = 0;

while workSheet.cell(row = rowEnd, column = binOutCol).value != None:
    rowEnd += 1
    numberOfTestcases += 1

#print(numberOfTestcases, rowEnd)

for row in range(rowStart, rowEnd):
    #Judge 1
    for column in range(maxErrorRateCol, maxErrorRateCol + 3):
        if workSheet.cell(row=row, column=column).value < constJudge1:
            judge1 = 'OK'
        else:
            judge1 = 'NG'
    
    #Judge 2
    judge2_1, judge2_2, judge2_3 = 0, 0, 0
    for column in range(countErrorRate1Col, countErrorRate1Col + 3):
        judge2_1 += workSheet.cell(row=row, column=column).value
    
    for column in range(countErrorRate2Col, countErrorRate2Col + 3):
        judge2_2 += workSheet.cell(row=row, column=column).value
    
    for column in range(countErrorRate3Col, countErrorRate3Col + 3):
        judge2_3 += workSheet.cell(row=row, column=column).value
    
    if judge1 == 'NG':
        if judge2_3 == 0:
            if judge2_2 == 0:
                if (math.ceil(workSheet.cell(row=row, column=binOutCol).value * constJudge2) * 3) >= judge2_1:
                    judge2_4 = 'OK'
                else:
                    judge2_4 = 'NG'
            else:
                judge2_4 = 'NG'
        else:
            judge2_4 = 'NG'
    else:
        judge2_4 = '-'
    
    #Judge by max error
    if judge2_4 == 'NG':
        for column in range(maxErrorCol, maxErrorCol + 2):
            if workSheet.cell(row=row, column=column).value <= constJudgeMaxError:
                judgeMaxError = 'OK'
            else:
                judgeMaxError = 'NG'
    else:
        judgeMaxError = '-'
    
    #Judge by bxp
    if (workSheet.cell(row=row, column=bxpOutCol).value - workSheet.cell(row=row, column=bxpCol).value) >= 2:
        for column in range(maxErrorRateCol, maxErrorRateCol + 3):
            if workSheet.cell(row=row, column=column).value < constJudgeBxp:
                judgeBxp = 'OK'
            else:
                judgeBxp = 'NG'
    else:
        judgeBxp = '-'
    #Final Judgement
    if (judge1 == 'OK') or (judge2_4 == 'OK') or (judgeMaxError == 'OK') or (judgeBxp == 'OK'):
        judgeFinal = 'OK'
        passed += 1
    else:
        judgeFinal = 'NG'
        failed += 1
    print(judge1, judge2_1, judge2_2, judge2_3, judge2_4, judgeMaxError, judgeBxp, judgeFinal)
print(' Passed %d case(s) \n Failed %d case(s)' %(passed, failed))
